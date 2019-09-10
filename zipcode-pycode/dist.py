import openpyxl
import sys
from geopy.distance import vincenty
import json

DIST_THRESHOLD = 200
ZIP = 'Zip'
LAT = 'Latitude'
LON = 'Longitude'
INPUT_FILE = 'zip-distance.xlsx'
SHEET = 'Sheet1'
OUTPUT_FILE = 'output.csv'

# Program starts from here
def main():

    mega_data_set = read_excel_data()

    data_list, zip_to_lat_long_map = store_data_in_mem(mega_data_set)
    #print (data_list)
    #print (zip_to_lat_long_map)

    zip_to_lat_long_map = calculate_zip_distances(data_list, zip_to_lat_long_map)
    print(json.dumps(zip_to_lat_long_map, indent=4, sort_keys=True))

    write_output_to_csv(zip_to_lat_long_map)


# This function reads our input excel sheet and stores 
# the excel data in python data map called 'mega_data_set'
def read_excel_data():

    mega_data_set = {}
    wb = openpyxl.load_workbook(INPUT_FILE)
    sheet = wb.get_sheet_by_name(SHEET)

    rows = sheet.max_row + 1
    columns = sheet.max_column + 1

    for observation in range(2, rows):
        data_set = {}
        for header in range(1, columns):
            data_set[str(sheet.cell(row=1, column=header).value)] = sheet.cell(row=observation, column=header).value

        mega_data_set[observation] = data_set

    return mega_data_set


# This function stores data read from excel sheet into a list and a python map
# List is required for iterating using 2 for loops; map is used
# to store the data associated with each zipcode which will finally
# be dumped into our output.csv file
def store_data_in_mem(mega_data_set):

    data_list = []
    tmp_list = []
    zip_to_lat_long_map = {}
    for _, value in mega_data_set.items():
        tmp_list.append(value[ZIP])
        tmp_list.append(value[LAT])
        tmp_list.append(value[LON])
        data_list.append(tmp_list)
        tmp_list = []
        if value[ZIP] not in zip_to_lat_long_map:
            zip_to_lat_long_map[str(value[ZIP])] = {LAT: value[LAT], LON: value[LON]}

    return sorted(data_list), zip_to_lat_long_map


# This function will iterate through our data list 
# and calculate distances between zipcodes using geopy
# It will also populate "0" and "1" as per zipcode distances
# into our zipcode map we crated above
def calculate_zip_distances(data_list, zip_to_lat_long_map):

    for i in range(0, len(data_list)):
        for j in range(i+1, len(data_list)):
            src = (data_list[i][1], data_list[i][2])
            dest = (data_list[j][1], data_list[j][2])
            src_new_col_entry = "zip_" + str(data_list[i][0])
            dest_new_col_entry = "zip_" + str(data_list[j][0])
            if src_new_col_entry not in zip_to_lat_long_map[data_list[i][0]]:
                zip_to_lat_long_map[data_list[i][0]].update({src_new_col_entry: 0})

            if dest_new_col_entry not in zip_to_lat_long_map[data_list[j][0]]:
                zip_to_lat_long_map[data_list[j][0]].update({dest_new_col_entry: 0})

            if (vincenty(src, dest).miles) < DIST_THRESHOLD:
                zip_to_lat_long_map[data_list[i][0]].update({dest_new_col_entry: 1})
                zip_to_lat_long_map[data_list[j][0]].update({src_new_col_entry: 1})
            else:
                zip_to_lat_long_map[data_list[i][0]].update({dest_new_col_entry: 0})
                zip_to_lat_long_map[data_list[j][0]].update({src_new_col_entry: 0})

    return zip_to_lat_long_map


# This function will write the final output from our map
# into an output.csv file
def write_output_to_csv(zip_to_lat_long_map):

    column_title = ZIP + "," + LAT + "," + LON
    for key, value in zip_to_lat_long_map.items():
        for item in sorted(value.keys()):
            if item == LAT or item == LON:
                continue
            else:
                column_title = column_title + "," + item

        break

    with open(OUTPUT_FILE, "a") as outfile:
        outfile.write(column_title + "\n")


    for key, value in zip_to_lat_long_map.items():
        entry = ""
        entry = key + "," + str(value[LAT]) + "," + str(value[LON])
        for item in sorted(value.keys()):
            if item == LAT or item == LON:
                continue
            else:
                entry = entry + "," + str(zip_to_lat_long_map[key][item])

        with open(OUTPUT_FILE, "a") as outfile:
            outfile.write(entry + "\n")

if __name__ == "__main__":
    main()
